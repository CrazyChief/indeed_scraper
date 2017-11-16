from tkinter import *
from datetime import datetime
from openpyxl import load_workbook
import jsonlines, xlsxwriter, proxy, scrape


parameters = ['q=', '&l=', '&jt=', '&radius=']
MAIN_LINK_PART = 'https://www.indeed.com/jobs?'
MAIN_FIELDS = ['Job Title', 'Job URL', 'Company', 'Company URL', 'City', 'State', 'Salary', 'Job Description']


def showSChoise():
    """Returns value of List or Manual"""
    value = sv.get()
    return value


def showLChoise():
    """Returns value of Use file or Keywords"""
    value = lv.get()
    return value


def showUseProxy():
    value = usePr.get()
    return value


def curr_time():
    """Returns date for naming file"""
    d = datetime.now()
    curr_date = str(d.year) + '_' + str(d.month) + '_' +\
                str(d.day) + '__' + str(d.hour) + '_' +\
                str(d.minute) + '_' + str(d.second)
    return curr_date


def read_states_from_excel():
    """Reads data from excel file and writes this data to .jsonl file.
    Returns file path."""
    wb = load_workbook(filename='states/Top50CitiesinUSwithZipCodes.xlsx')
    ws = wb.get_active_sheet()

    with jsonlines.open('states/tmpStates.jsonl', 'w') as f:
        for row in ws.rows:
            tmp_list = []
            for cell in row:
                tmp_list.append(str(cell.value))
            f.write(tmp_list)
        f.close()
    return 'states/tmpStates.jsonl'


def read_states_from_jsonl(filename):
    """Reads and returns data from .jsonl file.
    Returns list of data"""
    states = []
    with jsonlines.open(filename) as reader:
        for i, state in enumerate(reader):
            if (i == 0) or (i == 'None'):
                pass
            else:
                states.append(state)
        reader.close()
    return states


def build_link_part(first_par, last_par):
    """Creates part of link. Allowed types for last_par: list, string.
    Returns string."""
    result = ''.join(first_par)
    if type(last_par) == list:
        for lp in last_par:
            lp = re.sub(" ", "+", lp)
            result += ''.join(lp)
    elif type(last_par) == str:
        last_par = re.sub(" ", "+", last_par)
        result += ''.join(last_par)
    return result


def save_to_excel(data, name):
    """Save data to .xlsx file."""
    tm = curr_time()
    T.insert(END, "\n" + str(datetime.now().time()) +\
             ":\tStart saving data to /excel/" + tm + ".xlsx\n---------------------")

    workbook = xlsxwriter.Workbook('excel/' + tm + '.xlsx')
    worksheet = workbook.add_worksheet()

    worksheet.write('A1', MAIN_FIELDS[0])  # Job Title
    worksheet.write('B1', MAIN_FIELDS[1])  # Job URL
    worksheet.write('C1', MAIN_FIELDS[2])  # Company
    worksheet.write('D1', MAIN_FIELDS[3])  # Company URL
    worksheet.write('E1', MAIN_FIELDS[4])  # City
    worksheet.write('F1', MAIN_FIELDS[5])  # State
    worksheet.write('G1', MAIN_FIELDS[6])  # Salary
    worksheet.write('H1', MAIN_FIELDS[7])  # Job Description

    if len(data) != 0:
        for i, item in enumerate(data):
            try:
                worksheet.write('A' + str(i + 2), item['Job Title'])
            except KeyError:
                worksheet.write('A' + str(i + 2), '')
            try:
                worksheet.write('B' + str(i + 2), item['Job URL'])
            except KeyError:
                worksheet.write('B' + str(i + 2), '')
            try:
                worksheet.write('C' + str(i + 2), item['Company'])
            except KeyError:
                worksheet.write('C' + str(i + 2), '')
            try:
                worksheet.write('D' + str(i + 2), item['Company URL'])
            except KeyError:
                worksheet.write('D' + str(i + 2), '')
            try:
                worksheet.write('E' + str(i + 2), item['City'])
            except KeyError:
                worksheet.write('E' + str(i + 2), '')
            try:
                worksheet.write('F' + str(i + 2), item['State'])
            except KeyError:
                worksheet.write('F' + str(i + 2), '')
            try:
                worksheet.write('G' + str(i + 2), item['Salary'])
            except KeyError:
                worksheet.write('G' + str(i + 2), '')
            try:
                worksheet.write('H' + str(i + 2), item['Job Description'])
            except KeyError:
                worksheet.write('H' + str(i + 2), '')
    else:
        pass

    workbook.close()
    T.insert(END, "\n" + str(datetime.now().time()) + ":\tSaved\n---------------------")


def init_scrape():
    print(showUseProxy())
    if showUseProxy() == 1:
        userProxy = []
        userProxy.append(str(ip.get()))
        userProxy.append(str(port.get()))
        userProxy.append(str(typeSOCKS.get()))
    else:
        userProxy = 0
    if showSChoise() == 1:
        link = lf.get()
        # print(link)
        save_to_excel(scrape.parse(link, userProxy), "data")
    elif showSChoise() == 2:
        if kwr.get() != '':
            keywr = kwr.get()
            keywr = scrape.split_str(keywr, ";")
            kwr_part = build_link_part(parameters[0], keywr)
        else:
            kwr_part = ''

        if showLChoise() == 1:
            states = read_states_from_jsonl(read_states_from_excel())
            loc = []
            print(states)
            for st in states:
                print(st[0])
                loc.append(st[0])
        else:
            if location.get() != '':
                loc = location.get()
                loc = scrape.split_str(loc, "; ")
            else:
                loc = ''

        if area.get() != '':
            ar_part = build_link_part(parameters[3], area.get())
        else:
            ar_part = ''

        if terms.get() != '':
            term_part = build_link_part(parameters[2], terms.get())
        else:
            term_part = ''

        if loc != '':
            proxy_list = proxy.read_proxies(proxy.startCollectProxies())
            T.insert(END, "\n" + str(datetime.now().time()) + ":\tProxies had been collected." \
                                                              "\n---------------------")

            pcounter = 0
            for loc_part in loc:
                loc_part = build_link_part(parameters[1], loc_part)
                if pcounter == len(proxy_list) - 1:
                    pcounter = 0

                link = MAIN_LINK_PART + kwr_part + loc_part + ar_part + term_part
                T.insert(END, "\n" + str(datetime.now().time()) + ":\tStart scrapping data." \
                                                                  "\n---------------------")
                save_to_excel(scrape.parse(link, proxy_list[pcounter]), "Data")
        else:
            loc_part = ''
            link = MAIN_LINK_PART + kwr_part + loc_part + ar_part + term_part
            T.insert(END, "\n" + str(datetime.now().time()) + ":\tStart scrapping data." \
                                                              "\n---------------------")
            save_to_excel(scrape.parse(link, userProxy), "Data")

    else:
        T.insert(END, "\n" + str(datetime.now().time()) + ":\tYou didn\'t choice any option! (Link or Manual)" \
                 "\n---------------------")


if __name__ == '__main__':
    application = Tk()
    application.title('INDEED scrapper')
    application.geometry("1200x650+60+60")

    sv = IntVar()
    lv = IntVar()
    usePr = IntVar()

    Radiobutton(application,
                text="Link:",
                indicatoron=0,
                bg='darkgrey',
                padx=4,
                variable=sv,
                command=showSChoise,
                value=1).grid(row=0)
    Radiobutton(application,
                text="Manual:",
                indicatoron=0,
                bg='darkgrey',
                padx=4,
                variable=sv,
                command=showSChoise,
                value=2).grid(row=1)
    Label(application,
          text="Keywords: ",
          padx=15,
          pady=3).grid(row=2)
    Label(application,
          text="Radius: ",
          padx=15,
          pady=3).grid(row=3)
    Radiobutton(application,
                text="Use file:",
                indicatoron=0,
                bg='darkgrey',
                padx=4,
                variable=lv,
                command=showLChoise,
                value=1).grid(row=4)
    Radiobutton(application,
                text="City, state, or zip:",
                indicatoron=0,
                bg='darkgrey',
                padx=4,
                variable=lv,
                command=showLChoise,
                value=2).grid(row=4, column=1)
    Label(application,
          text="File path: /states/Top50CitiesinUSwithZipCodes.xlsx",
          bg="lightgrey",
          font="Times 12 italic",
          pady=3).grid(row=5)
    Label(application,
          text="Terms: ",
          pady=10).grid(row=6)

    lf = Entry(application, width=100)  # Link field
    kwr = Entry(application, width=80)  # Keywords field
    location = Entry(application, width=60)  # Location field
    area = Entry(application, width=20)  # Radius field
    area.insert(10, "100")  # Default radius value
    terms = Entry(application, width=30)    # Terms field
    lf.grid(row=0, column=1)
    kwr.grid(row=2, column=1)
    location.grid(row=5, column=1)
    area.grid(row=3, column=1)
    terms.grid(row=6, column=1)

    usep = Checkbutton(application,
                       text="Use Proxy",
                       variable=usePr).grid(row=7)
    Label(application,
          text="IP address:").grid(row=7, column=1)
    Label(application,
          text="Port").grid(row=8, column=1)
    Label(application,
          text="Type (4 or 5)").grid(row=9, column=1)
    ip = Entry(application, width=30)
    port = Entry(application, width=30)
    typeSOCKS = Entry(application, width=30)
    ip.grid(row=7, column=2)
    port.grid(row=8, column=2)
    typeSOCKS.grid(row=9, column=2)



    T = Text(application, height=20, width=120)
    T.grid(row=10, columnspan=4)
    quote = """This is indeed scrapper.\n---------------------"""
    T.insert(END, "\n" + quote)

    button = Button(application, text="Start scrap", width=24, fg="black", bg="lightgreen", pady=8, command=init_scrape)
    button.grid(row=11, column=1)

    application.mainloop()